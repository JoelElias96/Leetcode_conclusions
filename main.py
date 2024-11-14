from tkinter import *
from tkinter import messagebox as Massage
import pandas as pd
import openpyxl
from searchableCombobox import SearchableCombobox




   
   
def load_subjects():
        with open("Documents/subjectsList.txt", "r") as file:
            subjects = [line.strip() for line in file]
        subjects.sort()
        return subjects
    


def add_conclusion():
    subject = searchbox.get_selected_item()
    new_conclusion = conclusion_textbox.get("1.0", END).strip()

    # Load the existing workbook
    workbook = openpyxl.load_workbook("Documents/conclusionTable.xlsx")

    # Check if the sheet for the subject exists; if not, create it
    if subject not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=subject)
        sheet["A1"] = "Conclusion Text"
    else:
        sheet = workbook[subject]

    # Find the next empty row in the sheet
    next_row = sheet.max_row + 1

    # Add the new conclusion to the next row in Column A
    sheet.cell(row=next_row, column=1, value=new_conclusion)

    # Save the workbook
    workbook.save("Documents/conclusionTable.xlsx")
    conclusion_textbox.delete("1.0", END)
    Massage.showinfo("Success", f"Added conclusion to {subject}.")

def refresh_subjects(new_subject):
    searchbox.combo['values'] = load_subjects()
    searchbox.combo.set(new_subject)
    searchbox.update_options()
   


def edit_form():
    def save_new_subject():
        new_subject=entry_subject.get().capitalize()
        if Massage.askyesno("Adding new subject",f"Are you sure you want to add {new_subject} to the list of subjects?"):
            with open("Documents\subjectsList.txt", "a") as file:
                file.write(f"{new_subject}\n")     
        edit_window.destroy()
        refresh_subjects(new_subject)


    edit_window = Toplevel()
    edit_window.title("Edit format")
    edit_window.geometry("360x120")
    edit_window.configure(bg="white")
    edit_window.grab_set()
    edit_window.focus_set()
    edit_window.resizable(False, False)

    edit_label=Label(edit_window, text="Enter the subject you want to add:", font=("Arial", 10),background="white",pady=20,padx=10)
    edit_label.grid(row=0, column=0)

    #entry to add a new subject
    entry_subject=Entry(edit_window,highlightbackground="black",highlightthickness=2)
    entry_subject.grid(row=0, column=1)
    entry_subject.focus()

        
    #save button
    button_save = Button(edit_window, text="Save",command=save_new_subject,padx=2, pady=2)
    button_save.grid(row=1, column=0,padx=10, pady=10)

    #cancel button
    button_cancel= Button(edit_window, text="Cancel", command=edit_window.destroy,padx=2, pady=2)
    button_cancel.grid(row=1, column=1,padx=10, pady=10)

#window
window = Tk()
window.title("LeetCode Conclusions")
window.geometry("700x300")
window.configure(bg="white")
window.resizable(False, False)
icon = PhotoImage(file="images/leetcodeimage.png")
window.iconphoto(False, icon)


#leetcode image
leetcode_canva=Canvas(width=100, height=100,highlightthickness=0)
leetcode_image = PhotoImage(file="images/leetcodeimage.png")
image=leetcode_canva.create_image(50, 50, image=leetcode_image)
leetcode_canva.grid(column=0, row=0)

#main label
main_label = Label(window, text="LeetCode conclusions", font=("Arial", 40),background="white")
main_label.grid(row=0, column=1, columnspan=2)

#searchbox
searchbox=SearchableCombobox(window,load_subjects(),bg="white")
searchbox.grid(row=1, column=1)

 
#textbox for the conclusion
conclusion_textbox=Text(window, height=5, width=50, wrap=WORD, highlightbackground="black", highlightthickness=2)
searchbox.focus()
conclusion_textbox.grid(row=2, column=1)


#label for the name of the subject
label_subject = Label(window, text="Subject:",background="white")
label_subject.grid(row=1, column=0)

#label for the name of the takeaway
label_conclusion = Label(window, text="Conclusion:",background="white")
label_conclusion.grid(row=2, column=0)

   

#edit format button
link_edit=Button(window, text="Edit", command=edit_form)
link_edit.grid(row=3, column=0)

#button to add a new takeaway
button_add = Button(window, text="Add", command=add_conclusion,padx=2, pady=2)
button_add.grid(row=3, column=2,padx=10, pady=10)

#button to cancel
button_cancel= Button(window, text="Cancel", command=window.destroy,padx=2, pady=2)
button_cancel.grid(row=3, column=3,padx=10, pady=10)


window.mainloop()